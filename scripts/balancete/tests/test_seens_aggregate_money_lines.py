"""Subtotais só-numérico do PDF (Dom Felipe) não devem duplicar TOTAL no export Seens."""

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

import pytest

_SCRIPTS = Path(__file__).resolve().parents[1]


def _load_seens():
    path = _SCRIPTS / "export_excel_seens.py"
    spec = importlib.util.spec_from_file_location("export_excel_seens_agg", path)
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    return mod


@pytest.fixture(scope="module")
def seens():
    return _load_seens()


@pytest.mark.parametrize(
    "desc",
    [
        "1.549,40",
        "44,97",
        "15.284,86",
        "R$ 12.244,90",
        "r$ 12.244,90",
        "37.379,10 100,00",
        "20.818,90 55,70",  # só se a linha inteira for isso (sem texto)
    ],
)
def test_aggregate_money_or_value_percent_only(seens, desc: str):
    assert seens._is_money_only_line(desc) or seens._is_value_percent_only_line(desc)
    assert seens._is_aggregate_line(desc) is True


@pytest.mark.parametrize(
    "desc",
    [
        "Taxa de Condomínio 20.818,90 55,70",
        "Fundo de Reserva 2.081,89 5,57",
        "Seguro - P. 04/06 819,96 2,19",
        "Material de Limpeza - Condor - Esponja 42,36",
        "Taxa de Condomnio - Ref: 02/26 16.713,53",
    ],
)
def test_not_aggregate_real_descriptions(seens, desc: str):
    assert seens._is_aggregate_line(desc) is False


def test_dom_felipe_group_sum_no_duplicate_subtotal(seens, tmp_path):
    """Itens + linha só com subtotal (mesmo valor): subtotal não entra na soma."""
    data = {
        "metadata": {"condominio": "Dom Felipe", "competencia": "022026"},
        "entries": [
            {
                "section": "DESPESAS",
                "group": "Conservação e Manutenção",
                "descricao": "Serviço X 500,00",
                "valor": 500.0,
                "tipo_linha": "ITEM",
            },
            {
                "section": "DESPESAS",
                "group": "Conservação e Manutenção",
                "descricao": "Outro 500,00",
                "valor": 500.0,
                "tipo_linha": "ITEM",
            },
            {
                "section": "DESPESAS",
                "group": "Conservação e Manutenção",
                "descricao": "1.000,00",
                "valor": 1000.0,
                "tipo_linha": "ITEM",
            },
        ],
        "summary": {},
        "accounts": [],
    }
    out = tmp_path / "t.xlsx"
    seens.export_to_excel(data, str(out))
    from openpyxl import load_workbook

    wb = load_workbook(out, read_only=True, data_only=True)
    ws = wb.active
    detail_vals: list[float] = []
    total_val = None
    r = 11
    while r < 200:
        d = ws.cell(row=r, column=3).value
        v = ws.cell(row=r, column=12).value
        if d is None and v is None:
            break
        if isinstance(d, str) and d.strip().upper() == "TOTAL" and v is not None:
            total_val = float(v)
            break
        if v is not None and isinstance(v, (int, float)) and d and str(d).strip().upper() != "TOTAL":
            detail_vals.append(float(v))
        r += 1
    assert len(detail_vals) == 2
    assert detail_vals[0] + detail_vals[1] == pytest.approx(1000.0)
    assert total_val == pytest.approx(1000.0)
