"""
Preenche planilha XLSX a partir de JSON (openpyxl). Uso opcional / offline.

  python fill_template.py <template.xlsx> <saida.xlsx> <dados.json>

`dados.json` deve conter:
  { "rows": [ { "data", "descricao", "valor", "categoria", "grupo", "conta", "auditoria" } ] }

Se o template não tiver a aba esperada, cria planilha "Importacao_Balancete".
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.workbook import Workbook
except ImportError as e:  # pragma: no cover
    raise SystemExit(
        "openpyxl não instalado. Rode: pip install -r scripts/balancete/requirements.txt"
    ) from e

SHEET = "Importacao_Balancete"
HEADERS = [
    "Data",
    "Descrição",
    "Valor",
    "Categoria",
    "Grupo origem",
    "Conta (resumo)",
    "Linha original (auditoria)",
]


def main() -> None:
    if len(sys.argv) < 4:
        print(
            "Uso: python fill_template.py <template.xlsx> <saida.xlsx> <dados.json>",
            file=sys.stderr,
        )
        sys.exit(2)

    template_path = Path(sys.argv[1])
    out_path = Path(sys.argv[2])
    data_path = Path(sys.argv[3])

    payload = json.loads(data_path.read_text(encoding="utf-8"))
    rows = payload.get("rows") or []

    if template_path.exists():
        wb = load_workbook(template_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if SHEET in wb.sheetnames:
        ws = wb[SHEET]
    else:
        ws = wb.create_sheet(SHEET)
        for col, h in enumerate(HEADERS, start=1):
            ws.cell(row=1, column=col, value=h)

    start_row = ws.max_row + 1 if ws.max_row > 1 else 2
    if ws.max_row == 1:
        start_row = 2

    for i, row in enumerate(rows):
        r = start_row + i
        ws.cell(r, 1, value=row.get("data") or "")
        ws.cell(r, 2, value=row.get("descricao") or "")
        ws.cell(r, 3, value=row.get("valor"))
        ws.cell(r, 4, value=row.get("categoria") or "")
        ws.cell(r, 5, value=row.get("grupo") or "")
        ws.cell(r, 6, value=row.get("conta") or "")
        ws.cell(r, 7, value=row.get("auditoria") or "")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


if __name__ == "__main__":
    main()
