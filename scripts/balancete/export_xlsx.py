"""
Exporta balancete para .xlsx em formato padrão (sem template em models/).
Três abas: Lancamentos, Resumo, Contas.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Tuple, Union

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def _num(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _entries_list(data: Mapping[str, Any]) -> List[Dict[str, Any]]:
    e = data.get("entries")
    return list(e) if isinstance(e, list) else []


def _summary_dict(data: Mapping[str, Any]) -> Dict[str, Any]:
    s = data.get("summary")
    return dict(s) if isinstance(s, dict) else {}


def _accounts_list(data: Mapping[str, Any]) -> List[Dict[str, Any]]:
    a = data.get("accounts")
    return list(a) if isinstance(a, list) else []


def _entry_row(e: Mapping[str, Any]) -> Tuple[str, str, str, Optional[float], int]:
    sec = _str(e.get("section") or e.get("Section") or e.get("secao"))
    grp = _str(e.get("group") or e.get("Group") or e.get("grupo"))
    desc = _str(e.get("descricao") or e.get("Descricao") or e.get("description"))
    val = _num(e.get("valor") or e.get("Valor"))
    ord_raw = e.get("ordem")
    try:
        ordem = int(ord_raw) if ord_raw is not None else 0
    except (TypeError, ValueError):
        ordem = 0
    return sec, grp, desc, val, ordem


def _account_row(a: Mapping[str, Any]) -> Dict[str, Any]:
    return {
        "nome": _str(a.get("nome") or a.get("Nome") or a.get("name")),
        "saldo_anterior": _num(a.get("saldo_anterior")) or 0.0,
        "creditos": _num(a.get("creditos")) or 0.0,
        "debitos": _num(a.get("debitos")) or 0.0,
        "transferencias_mais": _num(a.get("transferencias_mais")) or 0.0,
        "transferencias_menos": _num(a.get("transferencias_menos")) or 0.0,
        "saldo_final": _num(a.get("saldo_final")) or 0.0,
    }


SUMMARY_ORDER: Tuple[str, ...] = (
    "total_receitas",
    "total_despesas",
    "saldo_anterior",
    "saldo_mes",
    "saldo_atual",
)


def _summary_rows(summary: Dict[str, Any]) -> List[Tuple[str, float]]:
    rows: List[Tuple[str, float]] = []
    seen = set()
    for key in SUMMARY_ORDER:
        if key not in summary:
            continue
        n = _num(summary[key])
        if n is None:
            continue
        rows.append((key, n))
        seen.add(key)
    for key in sorted(summary.keys()):
        if key in seen:
            continue
        n = _num(summary[key])
        if n is None:
            continue
        rows.append((key, float(n)))
    return rows


def export_balancete_to_xlsx(data: Union[Dict[str, Any], Mapping[str, Any]], output_path: str) -> str:
    """Cria workbook novo, três abas, formatação mínima. Retorna output_path."""
    if not isinstance(data, Mapping):
        data = {}

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    bold = Font(bold=True)
    dec_fmt = "#,##0.00"

    # --- Lancamentos ---
    ws_l = wb.create_sheet("Lancamentos", 0)
    h_l = ["Secao", "Grupo", "Descricao", "Valor", "Ordem"]
    for c, title in enumerate(h_l, start=1):
        cell = ws_l.cell(row=1, column=c, value=title)
        cell.font = bold
    entries = _entries_list(data)
    for r, ent in enumerate(entries, start=2):
        sec, grp, desc, val, ordem = _entry_row(ent if isinstance(ent, Mapping) else {})
        ws_l.cell(row=r, column=1, value=sec or None)
        ws_l.cell(row=r, column=2, value=grp or None)
        ws_l.cell(row=r, column=3, value=desc or None)
        vcell = ws_l.cell(row=r, column=4, value=val)
        if val is not None:
            vcell.number_format = dec_fmt
        ws_l.cell(row=r, column=5, value=ordem if ordem else r - 1)
    ws_l.column_dimensions["A"].width = 14
    ws_l.column_dimensions["B"].width = 28
    ws_l.column_dimensions["C"].width = 56
    ws_l.column_dimensions["D"].width = 16
    ws_l.column_dimensions["E"].width = 8

    # --- Resumo ---
    ws_r = wb.create_sheet("Resumo", 1)
    for c, title in enumerate(["Campo", "Valor"], start=1):
        cell = ws_r.cell(row=1, column=c, value=title)
        cell.font = bold
    summary = _summary_dict(data)
    pairs = _summary_rows(summary)
    for r, (campo, valor) in enumerate(pairs, start=2):
        ws_r.cell(row=r, column=1, value=campo)
        vc = ws_r.cell(row=r, column=2, value=valor)
        vc.number_format = dec_fmt
    ws_r.column_dimensions["A"].width = 22
    ws_r.column_dimensions["B"].width = 18

    # --- Contas ---
    ws_c = wb.create_sheet("Contas", 2)
    h_c = [
        "Nome",
        "Saldo Anterior",
        "Creditos",
        "Debitos",
        "Transferencias Mais",
        "Transferencias Menos",
        "Saldo Final",
    ]
    for c, title in enumerate(h_c, start=1):
        cell = ws_c.cell(row=1, column=c, value=title)
        cell.font = bold
    accounts = _accounts_list(data)
    for r, acc in enumerate(accounts, start=2):
        row = _account_row(acc if isinstance(acc, Mapping) else {})
        ws_c.cell(row=r, column=1, value=row["nome"] or None)
        for col, key in enumerate(
            [
                "saldo_anterior",
                "creditos",
                "debitos",
                "transferencias_mais",
                "transferencias_menos",
                "saldo_final",
            ],
            start=2,
        ):
            vc = ws_c.cell(row=r, column=col, value=row[key])
            vc.number_format = dec_fmt
    ws_c.column_dimensions["A"].width = 42
    for col in range(2, 8):
        ws_c.column_dimensions[get_column_letter(col)].width = 18

    wb.save(out)

    print("[INFO] Exportando balancete sem template", file=sys.stderr)
    print(f"[INFO] Total de lançamentos exportados: {len(entries)}", file=sys.stderr)
    print(f"[INFO] Total de contas exportadas: {len(accounts)}", file=sys.stderr)
    print(f"[INFO] Arquivo salvo em: {out.resolve()}", file=sys.stderr)

    return str(out.resolve())


def _load_json(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)
    return raw if isinstance(raw, dict) else {}


if __name__ == "__main__":
    sample = {
        "metadata": {"condominio": "Teste", "competencia": "02/2026"},
        "entries": [
            {
                "section": "RECEITAS",
                "group": "Receitas",
                "descricao": "Taxa de Condomínio",
                "valor": 1000.0,
                "ordem": 1,
            },
            {
                "section": "DESPESAS",
                "group": "Despesa com Pessoal",
                "descricao": "Salários",
                "valor": 500.0,
                "ordem": 2,
            },
        ],
        "summary": {
            "total_receitas": 1000.0,
            "total_despesas": 500.0,
            "saldo_anterior": 100.0,
            "saldo_mes": 500.0,
            "saldo_atual": 600.0,
        },
        "accounts": [
            {
                "nome": "Conta Movimento",
                "saldo_anterior": 100.0,
                "creditos": 1000.0,
                "debitos": 500.0,
                "transferencias_mais": 0.0,
                "transferencias_menos": 0.0,
                "saldo_final": 600.0,
            }
        ],
    }

    if len(sys.argv) >= 3:
        data_in = _load_json(sys.argv[1])
        export_balancete_to_xlsx(data_in, sys.argv[2])
    else:
        out_name = "teste_balancete.xlsx"
        if len(sys.argv) == 2:
            out_name = sys.argv[1]
        export_balancete_to_xlsx(sample, out_name)
