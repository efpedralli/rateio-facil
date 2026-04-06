"""
Exporta balancete canônico (mesmo JSON de `canonical_export.json`) para layout XLSX exigido pelo sistema Seens.
Usa apenas valores gravados nas células (sem fórmulas).
"""

from __future__ import annotations

import json
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Tuple, Union

from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side

DEC_FMT = "#,##0.00"
FONT_NAME = "Tahoma"
FONT_SIZE = 9

_THIN = Side(style="thin", color="FFAAAAAA")

_MESES = (
    "",
    "JANEIRO",
    "FEVEREIRO",
    "MARÇO",
    "ABRIL",
    "MAIO",
    "JUNHO",
    "JULHO",
    "AGOSTO",
    "SETEMBRO",
    "OUTUBRO",
    "NOVEMBRO",
    "DEZEMBRO",
)


def _strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def _norm_key(s: str) -> str:
    return _strip_accents(s.strip().lower())


def format_competencia(competencia: Union[int, str, None]) -> str:
    """
    Texto para A10 em CAIXA ALTA: COMPETÊNCIA FEVEREIRO DE 2026.
    Aceita: YYYYMM (int ou str), MM/YYYY, ISO periodo_inicio (YYYY-MM-DD).
    """
    ym = 0
    if competencia is None:
        return "COMPETÊNCIA —"
    if isinstance(competencia, int):
        ym = competencia if competencia > 0 else 0
    elif isinstance(competencia, str):
        t = competencia.strip()
        if re.fullmatch(r"\d{6}", t):
            ym = int(t)
        else:
            m = re.match(r"^(\d{1,2})\s*/\s*(\d{4})\s*$", t)
            if m:
                ym = int(m.group(2)) * 100 + int(m.group(1))
            else:
                m2 = re.search(r"(\d{4})-(\d{2})", t)
                if m2:
                    ym = int(m2.group(1)) * 100 + int(m2.group(2))
    if ym <= 0:
        return "COMPETÊNCIA —"
    year = ym // 100
    month = ym % 100
    if year < 1900 or year > 2100 or month < 1 or month > 12:
        return f"COMPETÊNCIA {ym}"
    return f"COMPETÊNCIA {_MESES[month]} DE {year}"


def _num(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _entries_list(data: Mapping[str, Any]) -> List[Dict[str, Any]]:
    e = data.get("entries")
    return list(e) if isinstance(e, list) else []


def _metadata_dict(data: Mapping[str, Any]) -> Dict[str, Any]:
    m = data.get("metadata")
    return dict(m) if isinstance(m, dict) else {}


def _summary_dict(data: Mapping[str, Any]) -> Dict[str, Any]:
    s = data.get("summary")
    return dict(s) if isinstance(s, dict) else {}


def _accounts_list(data: Mapping[str, Any]) -> List[Dict[str, Any]]:
    a = data.get("accounts")
    return list(a) if isinstance(a, list) else []


def _looks_like_transaction_line(text: str) -> bool:
    return bool(re.match(r"^\d{1,2}/\d{1,2}/\d{4}\s+", text.strip()))


def _infer_condominio_from_file_stem(stem: str) -> str:
    raw = re.split(r"[_\s-]+", stem)
    parts = [p for p in raw if p and not re.fullmatch(r"\d+", p)]
    base = " ".join(p.upper() for p in parts)
    if not base:
        return "CONDOMÍNIO"
    if not base.startswith("CONDOMÍNIO"):
        base = "CONDOMÍNIO " + base
    # Belle Chateau (modelo de validação do projeto)
    if re.search(r"BELLE.*CHATEAU|CHATEAU.*BELLE", base) and "RESIDENCIAL" not in base:
        base = base + " RESIDENCIAL"
    return base


def _resolve_condominio_display(data: Mapping[str, Any]) -> str:
    md = _metadata_dict(data)
    candidates = [
        md.get("condominio_nome"),
        md.get("condominio"),
        md.get("nome_condominio"),
        md.get("condominiumName"),
        md.get("condominium_name"),
    ]
    for c in candidates:
        s = _str(c)
        if not s:
            continue
        if _looks_like_transaction_line(s):
            continue
        return s.upper()

    for key in ("source_file", "sourceFile", "file_name", "fileName", "nome_arquivo"):
        p = _str(md.get(key))
        if not p:
            continue
        stem = Path(p.replace("\\", "/")).stem
        if stem:
            return _infer_condominio_from_file_stem(stem)

    # JSON bruto (parser) em metadata aninhado
    for key in ("raw_metadata", "parse_metadata"):
        sub = md.get(key)
        if isinstance(sub, dict):
            for sk in ("condominiumName", "condominio", "condominio_nome"):
                s = _str(sub.get(sk))
                if s and not _looks_like_transaction_line(s):
                    return s.upper()

    # Última tentativa: primeira linha de entries que pareça título (sem data, curta, sem R$)
    for ent in _entries_list(data):
        if not isinstance(ent, Mapping):
            continue
        d = _str(ent.get("descricao") or ent.get("Descricao") or ent.get("description"))
        if not d or _looks_like_transaction_line(d) or "R$" in d:
            continue
        if len(d) > 120:
            continue
        if re.search(r"data\s+fornecedor", d, re.I):
            continue
        if _norm_key(d) in ("receitas", "despesas"):
            continue
        return d.upper()

    return "CONDOMÍNIO"


def _infer_competencia_from_entries(entries: List[Dict[str, Any]]) -> int:
    for ent in entries:
        if not isinstance(ent, Mapping):
            continue
        d = _str(ent.get("descricao") or ent.get("Descricao") or ent.get("description"))
        m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})\s+", d)
        if m:
            mes, ano = int(m.group(2)), int(m.group(3))
            return ano * 100 + mes
    return 0


def _infer_competencia_value(data: Mapping[str, Any]) -> Union[int, str]:
    md = _metadata_dict(data)
    raw_ym = md.get("competencia_ym")
    if isinstance(raw_ym, int) and raw_ym > 0:
        return raw_ym
    if isinstance(raw_ym, str) and raw_ym.isdigit() and len(raw_ym) == 6:
        return int(raw_ym)

    label = _str(md.get("competencia"))
    if label:
        return label

    ini = _str(md.get("periodo_inicio") or md.get("periodoInicio"))
    if ini:
        return ini

    ym = _infer_competencia_from_entries(_entries_list(data))
    return ym if ym > 0 else 0


def _norm_section(raw: Any) -> str:
    s = _str(raw).upper().strip()
    if s in ("RECEITAS", "RECEITA"):
        return "RECEITAS"
    if s in ("DESPESAS", "DESPESA"):
        return "DESPESAS"
    return s


_GROUP_SKIP = frozenset(
    {
        "entradas valor",
        "saidas valor",
        "saídas valor",
        "receitas x despesas",
        "resumo_mes",
        "resumo mes",
    }
)


def _should_skip_group(sec: str, grp: str) -> bool:
    g = _norm_key(grp)
    if g in _GROUP_SKIP:
        return True
    # Encoding do PDF (ex.: SaÝdas → "saydas" após NFD)
    if re.match(r"^entradas\s+valor$", g):
        return True
    if re.match(r"^(saydas|saidas|sa[ií]das)\s+valor$", g):
        return True
    if g.startswith("resgates - fundo de obra"):
        return True
    if sec not in ("RECEITAS", "DESPESAS"):
        return True
    return False


def _entry_tipo_linha(ent: Mapping[str, Any]) -> str:
    for k in ("tipo_linha", "tipoLinha", "line_type", "lineType"):
        v = ent.get(k)
        if v is not None:
            return _str(v).upper()
    return ""


# Valor BR: 1.549,40 | 44,97 | 15.284,86
_RE_BR_DECIMAL = r"\d{1,3}(?:\.\d{3})*,\d{2}"

# Só valor monetário (subtotal colado do PDF, ex. Dom Felipe)
_RE_MONEY_ONLY_LINE = re.compile(
    rf"(?is)^\s*(?:R\s*\$|R\$|\$)?\s*\(?{_RE_BR_DECIMAL}\)?\s*$"
)

# Só valor + percentual (ex. total de grupo em receitas: 37.379,10 100,00)
_RE_VALUE_PERCENT_ONLY_LINE = re.compile(
    rf"(?is)^\s*{_RE_BR_DECIMAL}\s+{_RE_BR_DECIMAL}\s*$"
)


def _is_money_only_line(desc: str) -> bool:
    """Linha que é apenas um valor em formato BR (subtotal do PDF sem texto)."""
    t = re.sub(r"[\u00a0\s]+", " ", (desc or "").strip())
    return bool(t and _RE_MONEY_ONLY_LINE.match(t))


def _is_value_percent_only_line(desc: str) -> bool:
    """Linha só com dois números BR (valor e %); não confundir com 'Taxa ... 20.818,90 55,70'."""
    t = re.sub(r"[\u00a0\s]+", " ", (desc or "").strip())
    return bool(t and _RE_VALUE_PERCENT_ONLY_LINE.match(t))


def _is_aggregate_line(desc: str) -> bool:
    t = desc.strip()
    if not t:
        return True
    if _is_money_only_line(t):
        return True
    if _is_value_percent_only_line(t):
        return True
    low = _norm_key(t)

    if low in ("receitas", "despesas"):
        return True
    if re.match(r"^total\s*[:\.]", low):
        return True
    if re.search(r"\btotal\s+de\s+receitas\b", low):
        return True
    if re.search(r"\btotal\s+de\s+despesas\b", low):
        return True
    if re.match(r"^total\s*\(", low):
        return True
    if "subtotal" in low:
        return True
    if "total grupo" in low:
        return True
    if "total geral" in low:
        return True
    if "total de receitas" in low:
        return True
    if "total de despesas" in low:
        return True
    if "receitas do mes" in low or "receitas do mês" in low:
        return True
    if "despesas do mes" in low or "despesas do mês" in low:
        return True
    if re.match(r"^receitas\s+r\$", low):
        return True
    if re.match(r"^despesas\s+r\$", low):
        return True
    if re.match(r"^despesas\s+ordin", low):
        return True
    if "despesas" in low and "agua" in low and "esgoto" in low and "r$" in low:
        return True
    if re.match(r"^despesas\s+fundo\s+de\s+manuten", low):
        return True
    if "total dispon" in low:
        return True
    if "total (receitas - despesas)" in low:
        return True
    if re.match(r"^total\s*\(?\s*receitas\s*-\s*despesas", low):
        return True
    return False


_RE_GROUP_FROM_HEADER = re.compile(
    r"(?is)^\s*data\s+fornecedor\s*(\(\+\)|\(\-\))\s*(.+?)\s*valor\s*$"
)


def normalize_group_title(section: str, raw_group: str) -> str:
    """
    a) Belle / tabular: 'Data Fornecedor (+) Receitas Mensais Valor' → '(+) RECEITAS MENSAIS'.
    b) Layout simples (Dom Felipe, Dourados): 'Receitas', 'Conservação e Manutenção' →
       '(+) RECEITAS' / '(-) CONSERVAÇÃO E MANUTENÇÃO' conforme a seção ou o nome do grupo.
    """
    g = raw_group.strip()
    m = _RE_GROUP_FROM_HEADER.match(g)
    if m:
        sig, body = m.group(1), m.group(2).strip()
        body_u = re.sub(r"\s+", " ", body).upper()
        return f"{sig} {body_u}"
    collapsed = re.sub(r"\s+", " ", g).strip()
    gl = _norm_key(collapsed)
    revenue_name = gl.startswith("receita")
    if revenue_name or section == "RECEITAS":
        return f"(+) {collapsed.upper()}"
    if section == "DESPESAS":
        return f"(-) {collapsed.upper()}"
    return collapsed.upper()


_RE_DATE_LEAD = re.compile(r"^\d{1,2}/\d{1,2}/\d{4}\s+")
_RE_RS_TAIL = re.compile(r"\s+R\$\s*[\d\.\s\u00a0]*,\d{2}\s*$", re.IGNORECASE)
_RE_CONDOMINOS_PREFIX = re.compile(
    r"^(?:COND[OÔ]MINOS?\s+DO\s+CONDOM[IÍ]NIO\s+)",
    re.IGNORECASE,
)
# PDFs com encoding quebrado (ex.: CONDÈMINOS DO CONDOM═NIO) — parar na 1ª palavra após CONDOM
_RE_CONDOMINOS_PREFIX_FUZZY = re.compile(
    r"^.+?MINOS\s+DO\s+CONDOM[^\s]{1,16}\s+",
    re.IGNORECASE,
)
def _strip_company_prefix(t: str) -> str:
    """Remove razão social no início (ex.: 'QUIMIDROL ... LTDA MATERIAL...' -> 'MATERIAL...')."""
    m = re.match(
        r"^(.+?\b(?:LTDA\.?|S\.?\s*A\.?|S/A|EIRELI|EPP|ME))\s+",
        t,
        re.IGNORECASE,
    )
    if m and len(m.group(1).split()) >= 1:
        rest = t[m.end() :].strip()
        if len(rest) >= 8:
            return rest
    return t


_ENTITY_STRIP = (
    re.compile(r"^BANCO\s+COOPERATIVO\s+SICREDI\s+S\.?\s*A\.?\s+", re.I),
    re.compile(r"^RECEITA\s+FEDERAL\s+", re.I),
    re.compile(r"^DUPLIQUE\s+DESMBARGADOR\s+", re.I),
    re.compile(r"^DUPLIQUE\s+DESEMBARGADOR\s+", re.I),
    re.compile(r"^GEST\S+?\s+SERVI\S+?\s+ESPECIALIZADOS\s+LTDA\s+", re.I),
    re.compile(r"^ELEVADORES\s+ATLAS\s+SCHINDLER\s+", re.I),
    re.compile(r"^SANEPAR\s+-\s+CIA\s+DE\s+SANEAMENTO\s+", re.I),
    re.compile(r"^COPEL\s+-\s+CIA\s+PARAN\.\s+DE\s+ENERGIA\s+", re.I),
    re.compile(r"^NIO\s+FIBRA\s+\(OI\s+S/A\)\s+", re.I),
    re.compile(r"^EXTINCAMP\s+MANUT\.\s+DE\s+EXTINTORES\s+LTDA\s+", re.I),
    re.compile(r"^ANTONIO\s+APARECIDO\s+DA\s+SILVA\s+", re.I),
    re.compile(r"^PORTAS\s+MILENAR\s+", re.I),
    re.compile(r"^B\.M\.A\.\s+AUDITORIAS\s+CONDOMINIAIS\s+LTDA\s+", re.I),
    re.compile(r"^S[ÍI]NDICO\s+CARLOS\s+EDUARDO\s+NICOLEM\s+", re.I),
    re.compile(r"^SEGLINE\s+ADMINIST\.\s+DE\s+SERV\.\s+LTDA\s+", re.I),
    re.compile(r"^QUIMIDROL\s+COM\.\s+IND\.\s+IMPORTACAO\s+LTDA\s+", re.I),
)


def clean_item_description(desc: str) -> str:
    """Remove data inicial, fornecedor típico, sufixo R$; preserva núcleo da descrição."""
    t = desc.strip()
    if not t:
        return ""
    t = _RE_DATE_LEAD.sub("", t)
    t = _RE_RS_TAIL.sub("", t).strip()
    t = _RE_CONDOMINOS_PREFIX.sub("", t).strip()
    t = _RE_CONDOMINOS_PREFIX_FUZZY.sub("", t).strip()
    t = _strip_company_prefix(t)
    changed = True
    while changed:
        changed = False
        for rx in _ENTITY_STRIP:
            m = rx.match(t)
            if m:
                t = t[m.end() :].strip()
                changed = True
                break

    t = re.sub(r"\s+", " ", t).strip()
    return t.upper()


def _ordered_groups(
    entries: List[Dict[str, Any]],
) -> Tuple[List[Tuple[str, str]], Dict[Tuple[str, str], List[Dict[str, Any]]]]:
    order: List[Tuple[str, str]] = []
    groups: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    for ent in entries:
        if not isinstance(ent, Mapping):
            continue
        sec = _norm_section(
            ent.get("section")
            or ent.get("Section")
            or ent.get("secao")
            or ent.get("secaoMacro")
        )
        grp = (
            _str(
                ent.get("group")
                or ent.get("Group")
                or ent.get("grupo")
                or ent.get("grupoOrigem")
            )
            or "GERAL"
        )
        if _should_skip_group(sec, grp):
            continue
        key = (sec, grp)
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(dict(ent))
    receitas_keys = [k for k in order if k[0] == "RECEITAS"]
    despesas_keys = [k for k in order if k[0] == "DESPESAS"]
    return receitas_keys + despesas_keys, groups


def _thin_border() -> Border:
    return Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _apply_cell_style(cell: Any, *, bold: bool = False, number: bool = False) -> None:
    cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=bold)
    cell.border = _thin_border()
    if number:
        cell.number_format = DEC_FMT


def _money_round(x: float) -> float:
    return float(round(x + 1e-9, 2))


def _summary_totals(data: Mapping[str, Any], sum_rec: float, sum_desp: float) -> Tuple[float, float, float]:
    sm = _summary_dict(data)
    tr = _num(sm.get("total_receitas"))
    td = _num(sm.get("total_despesas"))
    sm_saldo = _num(sm.get("saldo_mes"))
    if tr is not None and td is not None:
        rec = _money_round(float(tr))
        desp = _money_round(float(td))
        if sm_saldo is not None:
            diff = _money_round(float(sm_saldo))
        else:
            diff = _money_round(rec - desp)
        return rec, desp, diff
    rec = _money_round(float(tr) if tr is not None else sum_rec)
    desp = _money_round(float(td) if td is not None else sum_desp)
    diff = _money_round(rec - desp)
    if sm_saldo is not None and abs(sm_saldo - diff) < 0.02:
        diff = _money_round(float(sm_saldo))
    return rec, desp, diff


def export_to_excel(data: dict, output_path: str) -> str:
    """
    Gera workbook Seens: cabeçalho fixo até linha 10, dados a partir da 11 (C=descrição, L=valor).
    Retorna o caminho absoluto salvo.
    """
    if not isinstance(data, Mapping):
        data = {}

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    md = _metadata_dict(data)
    condominio_display = _resolve_condominio_display(data)
    competencia_arg = _infer_competencia_value(data)

    wb = Workbook()
    ws = wb.active
    ws.title = "Balancete"

    for r in (5, 6, 7):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)

    ws["A8"] = condominio_display
    ws["A9"] = "BALANCETE MENSAL"
    ws["A10"] = format_competencia(competencia_arg)

    for addr in ("A8", "A9", "A10"):
        c = ws[addr]
        c.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
        c.border = _thin_border()

    entries = _entries_list(data)
    key_order, groups = _ordered_groups(entries)
    row = 11

    sum_receitas_grupos = 0.0
    sum_despesas_grupos = 0.0

    def write_section_header(text: str) -> None:
        nonlocal row
        cell = ws.cell(row=row, column=3, value=text)
        _apply_cell_style(cell, bold=True)
        row += 1

    for key in key_order:
        sec, grp = key
        block = groups.get(key, [])
        title = normalize_group_title(sec, grp)
        write_section_header(title)

        item_rows: List[Dict[str, Any]] = []
        for e in block:
            if not isinstance(e, Mapping):
                continue
            tl = _entry_tipo_linha(e)
            if tl in ("TOTAL_GERAL", "SUBTOTAL", "TITULO"):
                continue
            d = _str(e.get("descricao") or e.get("Descricao") or e.get("description"))
            if _is_aggregate_line(d):
                continue
            item_rows.append(e)

        total = 0.0
        for e in item_rows:
            raw_desc = _str(e.get("descricao") or e.get("Descricao") or e.get("description"))
            desc = clean_item_description(raw_desc)
            val = _num(e.get("valor") or e.get("Valor"))
            if val is None:
                val = 0.0
            total += val
            dc = ws.cell(row=row, column=3, value=desc or None)
            _apply_cell_style(dc)
            vc = ws.cell(row=row, column=12, value=_money_round(float(val)))
            _apply_cell_style(vc, number=True)
            row += 1

        total = _money_round(total)
        tc = ws.cell(row=row, column=3, value="TOTAL")
        _apply_cell_style(tc, bold=True)
        tcell = ws.cell(row=row, column=12, value=total)
        _apply_cell_style(tcell, bold=True, number=True)
        row += 1

        if sec == "RECEITAS":
            sum_receitas_grupos = _money_round(sum_receitas_grupos + total)
        else:
            sum_despesas_grupos = _money_round(sum_despesas_grupos + total)

    rec_fin, desp_fin, diff_fin = _summary_totals(data, sum_receitas_grupos, sum_despesas_grupos)

    lc = ws.cell(row=row, column=3, value="TOTAL RECEITAS DO MÊS")
    _apply_cell_style(lc, bold=True)
    vc = ws.cell(row=row, column=12, value=rec_fin)
    _apply_cell_style(vc, bold=True, number=True)
    row += 1

    lc = ws.cell(row=row, column=3, value="TOTAL DESPESAS DO MÊS")
    _apply_cell_style(lc, bold=True)
    vc = ws.cell(row=row, column=12, value=desp_fin)
    _apply_cell_style(vc, bold=True, number=True)
    row += 1

    hc = ws.cell(row=row, column=3, value="RESUMO DO MÊS")
    _apply_cell_style(hc, bold=True)
    row += 1

    for label, val in (
        ("RECEITAS", rec_fin),
        ("DESPESAS", desp_fin),
        ("TOTAL (RECEITAS - DESPESAS)", diff_fin),
    ):
        lc = ws.cell(row=row, column=3, value=label)
        _apply_cell_style(lc, bold=label.startswith("TOTAL"))
        vc = ws.cell(row=row, column=12, value=_money_round(float(val)))
        _apply_cell_style(vc, bold=label.startswith("TOTAL"), number=True)
        row += 1

    accounts = _accounts_list(data)
    if accounts:
        write_section_header(
            "RESUMO DAS CONTAS - POSIÇÃO CONSOLIDADA DA CONTA PESSOA JURÍDICA - SICREDI"
        )
        for acc in accounts:
            if not isinstance(acc, Mapping):
                continue
            nome = _str(acc.get("nome") or acc.get("Nome") or acc.get("name")) or "CONTA"
            write_section_header(nome.upper())
            pairs = [
                ("SALDO ANTERIOR", _num(acc.get("saldo_anterior")) or 0.0),
                ("ENTRADAS", _num(acc.get("creditos")) or 0.0),
                ("SAÍDAS", _num(acc.get("debitos")) or 0.0),
                ("SALDO ATUAL", _num(acc.get("saldo_final")) or 0.0),
            ]
            for lab, v in pairs:
                lc = ws.cell(row=row, column=3, value=lab)
                _apply_cell_style(lc)
                vc = ws.cell(row=row, column=12, value=_money_round(float(v)))
                _apply_cell_style(vc, number=True)
                row += 1

    ws.column_dimensions["C"].width = 72
    ws.column_dimensions["L"].width = 16

    wb.save(out)

    print("[INFO] Export Seens (openpyxl)", file=sys.stderr)
    print(f"[INFO] Linha inicial dados: 11 | grupos: {len(key_order)}", file=sys.stderr)
    print(f"[INFO] Arquivo salvo em: {out.resolve()}", file=sys.stderr)

    return str(out.resolve())


def _load_json(path: str) -> Dict[str, Any]:
    blob = Path(path).read_bytes()
    for enc in ("utf-8-sig", "utf-8", "utf-16", "utf-16-le", "utf-16-be", "cp1252"):
        try:
            text = blob.decode(enc)
            raw = json.loads(text)
            return raw if isinstance(raw, dict) else {}
        except (UnicodeDecodeError, json.JSONDecodeError):
            continue
    raise ValueError(f"JSON inválido ou encoding não suportado: {path}")


_FUND_CONTA_NOME = {
    "FUNDO_DE_OBRA": "FUNDO DE OBRA",
    "FUNDO_RESERVA_POUPANCA": "FUNDO DE RESERVA DE POUPANÇA PERMANENTE",
    "CONTA_CORRENTE_FLUXO": "CONTA CORRENTE - FLUXO DE CAIXA DE MANUTENÇÃO",
}

def _is_synthetic_resumo_line(desc: str) -> bool:
    low = _norm_key(desc or "")
    if not low:
        return True
    if low in ("receitas", "despesas"):
        return True
    if low.startswith("receitas r$"):
        return True
    if low.startswith("despesas r$"):
        return True
    if "total (receitas - despesas)" in low:
        return True
    if "total dispon" in low:
        return True
    if "saldo atual" in low and "fundo" in low:
        return True
    if "saldo atual" in low and "conta corrente" in low:
        return True
    if "saldo atual" in low and "fluxo" in low:
        return True
    if re.match(r"^despesas\s+ordin", low):
        return True
    if "despesas" in low and "agua" in low and "esgoto" in low:
        return True
    if re.match(r"^despesas\s+fundo\s+de\s+manuten", low):
        return True
    if "receitas do mes" in low and "total geral" in low:
        return True
    if "despesas do mes" in low and "total geral" in low:
        return True
    if "total grupo" in low or "total geral" in low or "subtotal" in low:
        return True
    if re.search(r"\btotal\s+de\s+receitas\b", low):
        return True
    if re.search(r"\btotal\s+de\s+despesas\b", low):
        return True
    if re.match(r"^total\s*[:\.]", low):
        return True
    if re.match(r"^total\s+\(", low):
        return True
    return False


def _should_skip_export_group_belle(grp: str) -> bool:
    g = _norm_key(grp or "")
    if not g:
        return False
    if g in ("resumo_mes", "resumo mes", "receitas x despesas"):
        return True
    if g.startswith("resgates - fundo de obra"):
        return True
    if g == "entradas valor":
        return True
    if re.match(r"^(saydas|saidas|sa[ií]das)\s+valor$", g):
        return True
    return False


def _fund_from_resumo_conta_desc(desc: str) -> Optional[str]:
    low = _norm_key(desc or "")
    if "total dispon" in low:
        return None
    if "fundo de reserva" in low or ("reserva" in low and "poup" in low):
        return "FUNDO_RESERVA_POUPANCA"
    if "fundo de obra" in low:
        return "FUNDO_DE_OBRA"
    if (
        "conta corrente" in low
        or "fluxo de caixa de manuten" in low
        or "fluxo de caixa de manutencao" in low
        or "receita de cotas" in low
        or ("receita mensal" in low and "fluxo" in low)
        or "aluguel salao" in low
        or ("aluguel" in low and "salao" in low)
    ):
        return "CONTA_CORRENTE_FLUXO"
    return None


def _fund_from_saldo_atual_desc(desc: str) -> Optional[str]:
    low = _norm_key(desc or "")
    if "fundo de obra" in low and "reserva" not in low:
        return "FUNDO_DE_OBRA"
    if "reserva" in low and "poup" in low:
        return "FUNDO_RESERVA_POUPANCA"
    if "conta corrente" in low or "fluxo de caixa" in low:
        return "CONTA_CORRENTE_FLUXO"
    return None


def _pivot_accounts_from_resumo_belle(
    raw: Dict[str, Any], summary: Dict[str, float]
) -> List[Dict[str, Any]]:
    rows = raw.get("resumoContas") or []
    if not isinstance(rows, list) or not rows:
        return []
    resumo_rows_classified = 0
    aggs: Dict[str, Dict[str, float]] = {
        k: {
            "nome": _FUND_CONTA_NOME[k],
            "saldo_anterior": 0.0,
            "creditos": 0.0,
            "debitos": 0.0,
            "transferencias_mais": 0.0,
            "transferencias_menos": 0.0,
            "saldo_final": 0.0,
        }
        for k in _FUND_CONTA_NOME
    }
    for r in rows:
        if not isinstance(r, Mapping):
            continue
        desc = _str(r.get("descricao"))
        fk = _fund_from_resumo_conta_desc(desc)
        if not fk:
            continue
        resumo_rows_classified += 1
        a = aggs[fk]
        low = _norm_key(desc)
        v = float(abs(_num(r.get("valor")) or 0.0))
        mov = _str(r.get("movimento")).upper()
        if mov == "SALDO_ANTERIOR" or (
            "acumulado" in low
            and (
                "anterior" in low or "competencia" in low or "competência" in low
            )
        ):
            a["saldo_anterior"] = v
        elif mov == "SALDO_ATUAL" or "saldo atual" in low:
            a["saldo_final"] = v
        elif mov == "SAIDA" or (
            fk == "FUNDO_RESERVA_POUPANCA" and "resgate" in low
        ):
            a["debitos"] += v
        elif mov in ("ENTRADA", "TOTAL_DISPONIVEL"):
            a["creditos"] += v

    if resumo_rows_classified == 0:
        return []

    for e in raw.get("entries") or []:
        if not isinstance(e, Mapping):
            continue
        if _str(e.get("fase")) != "RESUMO_MES":
            continue
        desc = _str(e.get("descricao"))
        low = _norm_key(desc)
        if "saldo atual" not in low:
            continue
        fk = _fund_from_saldo_atual_desc(desc)
        if not fk:
            continue
        val = _num(e.get("valor"))
        if val is not None:
            aggs[fk]["saldo_final"] = float(abs(val))

    tr = _num(summary.get("total_receitas"))
    td = _num(summary.get("total_despesas"))
    obra = aggs["FUNDO_DE_OBRA"]
    res = aggs["FUNDO_RESERVA_POUPANCA"]
    flux = aggs["CONTA_CORRENTE_FLUXO"]
    looks_like_multi_fund = (
        obra["saldo_anterior"] > 0.01
        or res["saldo_anterior"] > 0.01
        or (obra["creditos"] > 0.01 and res["creditos"] > 0.01)
    )
    if tr is not None and td is not None and looks_like_multi_fund:
        flux["creditos"] = _money_round(float(tr) - obra["creditos"] - res["creditos"])
        flux["debitos"] = _money_round(float(td))

    order = (
        "FUNDO_DE_OBRA",
        "FUNDO_RESERVA_POUPANCA",
        "CONTA_CORRENTE_FLUXO",
    )
    return [aggs[k] for k in order]


def _canonical_export_from_saida_belle(raw: Dict[str, Any]) -> Dict[str, Any]:
    """Monta payload tipo canonical_export a partir de JSON schemaVersion 2 (teste manual)."""
    if raw.get("schemaVersion") != 2:
        return raw
    md_in = raw.get("metadata") if isinstance(raw.get("metadata"), dict) else {}
    entries_out: List[Dict[str, Any]] = []
    for e in raw.get("entries") or []:
        if not isinstance(e, Mapping):
            continue
        fase = e.get("fase")
        if fase not in (None, "LANCAMENTOS"):
            continue
        tl = _str(e.get("tipoLinha") or e.get("tipo_linha")).upper()
        if tl != "ITEM":
            continue
        sec = e.get("secaoMacro") or e.get("section")
        if sec not in ("RECEITAS", "DESPESAS"):
            continue
        desc = _str(e.get("descricao") or e.get("description"))
        grp = e.get("grupoOrigem") or e.get("group") or ""
        if _should_skip_export_group_belle(_str(grp)):
            continue
        if _is_synthetic_resumo_line(desc):
            continue
        entries_out.append(
            {
                "section": sec,
                "group": e.get("grupoOrigem") or e.get("group"),
                "descricao": e.get("descricao") or e.get("description"),
                "valor": e.get("valor"),
                "tipo_linha": e.get("tipoLinha") or e.get("tipo_linha"),
            }
        )
    canon = raw.get("canonical") or {}
    resumo = canon.get("resumo") if isinstance(canon.get("resumo"), list) else []
    summary: Dict[str, float] = {}
    for item in resumo:
        if not isinstance(item, Mapping):
            continue
        lab = _norm_key(_str(item.get("label")))
        val = _num(item.get("valor"))
        if val is None:
            continue
        if "receita" in lab and "total" in lab and "geral" in lab and "despes" not in lab:
            summary["total_receitas"] = float(val)
        elif "despes" in lab and "total" in lab and "geral" in lab:
            summary["total_despesas"] = float(val)

    for e in raw.get("entries") or []:
        if not isinstance(e, Mapping):
            continue
        if e.get("fase") not in (None, "LANCAMENTOS"):
            continue
        low = _norm_key(_str(e.get("descricao")))
        val = _num(e.get("valor"))
        if val is None:
            continue
        if (
            "receitas" in low
            and "total" in low
            and "geral" in low
            and "despes" not in low
        ):
            summary["total_receitas"] = float(val)
        if "despesas" in low and "total" in low and "geral" in low:
            summary["total_despesas"] = float(val)
        if re.search(r"\btotal\s+de\s+receitas\b", low):
            summary["total_receitas"] = float(val)
        if re.search(r"\btotal\s+de\s+despesas\b", low):
            summary["total_despesas"] = float(val)
        if "total" in low and "receitas" in low and "despesas" in low and "-" in low:
            summary["saldo_mes"] = float(abs(val))

    accounts = _pivot_accounts_from_resumo_belle(raw, summary)

    meta_out = {
        "condominio": md_in.get("condominiumName") or md_in.get("condominio"),
        "competencia": md_in.get("competenceLabel"),
        "periodo_inicio": None,
        "source_file": md_in.get("fileName") or md_in.get("source_file"),
        "parser_type": md_in.get("parserLayoutId"),
    }
    return {
        "metadata": meta_out,
        "entries": entries_out,
        "summary": summary,
        "accounts": accounts,
    }


if __name__ == "__main__":
    sample = {
        "metadata": {
            "condominio": "Residencial Exemplo",
            "competencia": "04/2026",
            "periodo_inicio": "2026-04-01",
            "periodo_fim": "2026-04-30",
        },
        "entries": [
            {
                "section": "RECEITAS",
                "group": "Data Fornecedor (+) Receitas Mensais Valor",
                "descricao": "Taxa condominial",
                "valor": 15000.0,
                "ordem": 1,
            },
            {
                "section": "DESPESAS",
                "group": "Data Fornecedor (-) Despesas Diversas Valor",
                "descricao": "Salários",
                "valor": 8000.0,
                "ordem": 2,
            },
        ],
        "summary": {"total_receitas": 15000.0, "total_despesas": 8000.0, "saldo_mes": 7000.0},
        "accounts": [],
    }

    if len(sys.argv) >= 3:
        payload = _load_json(sys.argv[1])
        if payload.get("schemaVersion") == 2 and "canonical" in payload:
            payload = _canonical_export_from_saida_belle(payload)
        export_to_excel(payload, sys.argv[2])
    elif len(sys.argv) == 2 and Path(sys.argv[1]).exists():
        payload = _load_json(sys.argv[1])
        if payload.get("schemaVersion") == 2 and "canonical" in payload:
            payload = _canonical_export_from_saida_belle(payload)
        out_default = Path("outputs") / "cli_seens_out.xlsx"
        export_to_excel(payload, str(out_default))
        print(f"Gerado: {out_default.resolve()}", file=sys.stderr)
    else:
        out_default = Path("outputs") / "Residencial_Exemplo_202604_seens.xlsx"
        export_to_excel(sample, str(out_default))
        print(f"Exemplo salvo em: {out_default.resolve()}", file=sys.stderr)
